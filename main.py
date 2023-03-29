# -*- coding: UTF-8 -*-
import os
import os.path
import logging
import logging.config
import sys
import configparser
import time
import shutil
import openpyxl                       # Для .xlsx
#import xlrd                          # для .xls
from   price_tools import getCellXlsx, getCell, nameToId, currencyTypeX, sheetByName
import csv
import urllib.request



def getXlsString(sh, i, in_columns_j):
    impValues = {}
    for item in in_columns_j.keys() :
        j = in_columns_j[item]-1
        if item in ('закупка','продажа','цена1') :
            if (getCell(row=i, col=j, isDigit='N', sheet=sh).find('По запросу') >=0 or
               getCell(row=i, col=j, isDigit='N', sheet=sh).find('стоимости') >=0):
                impValues[item] = '0.1'
            else :
                impValues[item] = getCell(row=i, col=j, isDigit='Y', sheet=sh)
            #print(sh, i, sh.cell( row=i, column=j).value, sh.cell(row=i, column=j).number_format, currencyType(sh, i, j))
        elif item == 'валюта_по_формату':
            impValues[item] = currencyType(row=i, col=j, sheet=sh)
        else:
            impValues[item] = getCell(row=i, col=j, isDigit='N', sheet=sh)
    return impValues



def getXlsxString(sh, i, in_columns_j):
    impValues = {}
    for item in in_columns_j.keys() :
        j = in_columns_j[item]
        if item in ('закупка','продажа','цена2','цена1') :
            sss = getCellXlsx(row=i, col=j, isDigit='N', sheet=sh)
            if (sss.find('запросу') >=0 or sss.find('стоимости') >= 0):
                impValues[item] = '0.1'
            else :
                impValues[item] = getCellXlsx(row=i, col=j, isDigit='Y', sheet=sh)
            #print(sh, i, sh.cell( row=i, column=j).value, sh.cell(row=i, column=j).number_format, currencyType(sh, i, j))
        elif item == 'валюта_по_формату':
            #impValues[item] = 'RUR'
            impValues[item] = currencyTypeX(row=i, col=j, sheet=sh)
        else:
            impValues[item] = getCellXlsx(row=i, col=j, isDigit='N', sheet=sh)
    return impValues



def convert_excel2csv(cfg):
    priceFName= cfg.get('basic','filename_in')
    sheetName = cfg.get('basic','sheetname')


    log.debug('Reading file ' + priceFName )
    book, sheet = sheetByName(fileName = priceFName, sheetName = sheetName)
    if not sheet:
        log.error("Нет листа "+sheetName+" в файле "+ priceFName)
        return False
    log.debug("Sheet   "+sheetName)
    out_cols = cfg.options("cols_out")
    in_cols  = cfg.options("cols_in")
    out_template = {}
    for vName in out_cols :
         out_template[vName] = cfg.get("cols_out", vName)
    in_cols_j = {}
    for vName in in_cols :
         in_cols_j[vName] = cfg.getint("cols_in",  vName)
    #brands,   discount     = config_read(cfgFName, 'discount')
    #for k in discount.keys():
    #    discount[k] = (100 - int(discount[k]))/100
    #print(discount)

    outFileUSD = False
    outFileEUR = False
    outFileRUR = False
    if cfg.has_option('basic','filename_out_RUR'):
        csvFfileNameRUR = cfg.get('basic', 'filename_out_RUR')
        outFileRUR = open(csvFfileNameRUR, 'w', newline='')
        csvWriterRUR = csv.DictWriter(outFileRUR, fieldnames=cfg.options('cols_out'))
        csvWriterRUR.writeheader()
    if cfg.has_option('basic', 'filename_out_USD'):
        csvFfileNameUSD = cfg.get('basic', 'filename_out_USD')
        outFileUSD = open(csvFfileNameUSD, 'w', newline='')
        csvWriterUSD = csv.DictWriter(outFileUSD, fieldnames=cfg.options('cols_out'))
        csvWriterUSD.writeheader()
    if cfg.has_option('basic', 'filename_out_EUR'):
        csvFfileNameEUR = cfg.get('basic', 'filename_out_EUR')
        outFileEUR = open(csvFfileNameEUR, 'w', newline='')
        csvWriterEUR = csv.DictWriter(outFileEUR, fieldnames=cfg.options('cols_out'))
        csvWriterEUR.writeheader()


    '''                                     # Блок проверки свойств для распознавания групп      XLSX
    for i in range(2, 15):
        i_last = i
        ccc = sheet.cell( row=i, column=in_cols_j['подгруппа'] )
        print(i, sheet.cell(row=i, column=in_cols_j['цена1']).value, ccc.value)
        print(ccc.font.name, ccc.font.sz, ccc.font.b, ccc.font.i, '------', 'ccc.font.color.rgb', ccc.fill.bgColor.rgb, 'ccc.fill.fgColor.rgb')
        print('------')
    return
    '''
    '''                                     # Блок проверки свойств для распознавания групп      XLS                                  
    for i in range(19, 25):                                                         
        xfx = sheet.cell_xf_index(i, 1)
        xf  = book.xf_list[xfx]
        bgci  = xf.background.pattern_colour_index
        fonti = xf.font_index
        ccc = sheet.cell(i, 1)
        if ccc.value == None :
            print (i, colSGrp, 'Пусто!!!')
            continue
                                         # Атрибуты шрифта для настройки конфига
        font = book.font_list[fonti]
        print( '---------------------- Строка', i, '-----------------------', sheet.cell(i, 1).value)
        print( 'background_colour_index=',bgci)
        print( 'fonti=', fonti, '           xf.alignment.indent_level=', xf.alignment.indent_level)
        print( 'bold=', font.bold)
        print( 'weight=', font.weight)
        print( 'height=', font.height)
        print( 'italic=', font.italic)
        print( 'colour_index=', font.colour_index )
        print( 'name=', font.name)
    return
    '''

    recOut  ={}
    subgrp = ''
    grp = ''
    for i in range(1, sheet.max_row +1) :                               # xlsx
#   for i in range(1, sheet.nrows) :                                     # xls
        i_last = i
        try:
#            xfx = sheet.cell_xf_index(i, 1)                              # xls
#            xf  = book.xf_list[xfx]                                      # xls
#            bgci  = xf.background.pattern_colour_index                   # xls
            impValues = getXlsxString(sheet, i, in_cols_j)                # xlsx
            #impValues = getXlsString(sheet, i, in_cols_j)                # xls
            #print( impValues )
            ccc1 = sheet.cell(row=i, column=in_cols_j['цена1']).value

            if sheetName in ('VS', 'CO', 'PAVA'):
                if (sheet.cell(row=i, column=in_cols_j['подгруппа']).font.b is True and
                    sheet.cell(row=i, column=in_cols_j['цена1']).value is None):          # подгруппа
                    subgrp = impValues['подгруппа']
                    continue
                elif (impValues['код_'] == '' or
                    impValues['код_'] == 'SAP' or
                    impValues['цена1'] == '0'):                                           # лишняя строка
                    continue
                impValues['подгруппа'] = subgrp
                impValues['описание'] = impValues['описание'].encode('cp1251', errors='replace').decode('cp1251')

            else:
                log.error('нераспознан sheetName "%s"', sheetName)      # далее общая для всех обработка

            for outColName in out_template.keys() :
                shablon = out_template[outColName]
                for key in impValues.keys():
                    if shablon.find(key) >= 0:
                        shablon = shablon.replace(key, impValues[key])
                if (outColName == 'закупка') and ('*' in shablon) :
                    p = shablon.find("*")
                    vvv1 = float(shablon[:p])
                    vvv2 = float(shablon[p+1:])
                    shablon = str(round(vvv1 * vvv2, 2))
                recOut[outColName] = shablon.strip()

            recOut['код'] = nameToId(recOut['код'])
            if  recOut['продажа'] == '0.1':
                recOut['валюта'] = 'USD'
                recOut['закупка'] = '0.1'
            if recOut['валюта'] == 'RUR':
                csvWriterRUR.writerow(recOut)
            elif recOut['валюта'] == 'USD':
                csvWriterUSD.writerow(recOut)
            elif recOut['валюта'] == 'EUR':
                csvWriterEUR.writerow(recOut)
            else:
                log.error('нераспознана валюта "%s" для товара "%s"', recOut['валюта'], recOut['код производителя'])

        except Exception as e:
            print(e)
            if str(e) == "'NoneType' object has no attribute 'rgb'":
                pass
            else:
                log.debug('Exception: <' + str(e) + '> при обработке строки ' + str(i) + '.')

    log.info('Обработано ' + str(i_last) + ' строк.')
    if outFileRUR:
        outFileRUR.close()
    if outFileUSD:
        outFileUSD.close()
    if outFileEUR:
        outFileEUR.close()



def download( cfg ):
    retCode     = False
    filename_new= cfg.get('basic','filename_new')
    filename_old= cfg.get('basic','filename_old')
    url_file = cfg.get('download','url_file')

    filePrice = u'specvideoproject.xlsx'

    try:
        sss = urllib.request.urlopen(url_file).read()       # Скачиваем сначала страницу
        log.info('Размер скачанного файла %s байе', len(sss))
        if os.path.exists(filePrice):
            os.remove(filePrice)
        f = open(filePrice, 'wb')  # Теперь записываем файл
        f.write(sss)
        f.close()
        if os.path.exists(filename_new):
            if os.path.exists(filename_old):
                os.remove(filename_old)
            os.rename(filename_new, filename_old)
        os.rename(filePrice, filename_new)
        retCode = True
    except Exception as e:
        log.debug('Exception: <' + str(e) + '>')

    return retCode




def config_read( cfgFName ):
    cfg = configparser.ConfigParser(inline_comment_prefixes=('#'))
    if  os.path.exists('private.cfg'):
        cfg.read('private.cfg', encoding='utf-8')
    if  os.path.exists('getting.cfg'):
        cfg.read('getting.cfg', encoding='utf-8')
    if  os.path.exists(cfgFName):
        cfg.read( cfgFName, encoding='utf-8')
    else:
        log.debug('Нет файла конфигурации '+cfgFName)
    return cfg



def is_file_fresh(fileName, qty_days):
    qty_seconds = qty_days *24*60*60
    if os.path.exists( fileName):
        price_datetime = os.path.getmtime(fileName)
    else:
        log.error('Не найден файл  '+ fileName)
        return False

    file_age = round((time.time() - price_datetime) / 24 / 60 / 60)
    if file_age > qty_days :
        log.error('Файл "' + fileName + '" устарел! Допустимый период ' + str(qty_days)+' дней, а ему ' + str(file_age))
        return False
    else:
        return True



def make_loger():
    global log
    logging.config.fileConfig('logging.cfg')
    log = logging.getLogger('logFile')


def main(dealerName):
    """ Обработка прайсов выполняется согласно файлов конфигурации.
    Для этого в текущей папке должны быть файлы конфигурации, описывающие
    свойства файла и правила обработки. По одному конфигу на каждый
    прайс или раздел прайса со своими правилами обработки
    """
    make_loger()
    log.info('          ' + dealerName)

    rc_download = False
    '''
    '''
    if os.path.exists('getting.cfg'):
        cfg = config_read('getting.cfg')
        filename_new = cfg.get('basic','filename_new')
        if cfg.has_section('download'):
            rc_download = download(cfg)
        if not(rc_download==True or is_file_fresh( filename_new, int(cfg.get('basic','срок годности')))):
            return False
    for cfgFName in os.listdir("."):
        if cfgFName.startswith("cfg") and cfgFName.endswith(".cfg"):
            log.info('----------------------- Processing '+cfgFName )
            cfg = config_read(cfgFName)
            filename_in = cfg.get('basic','filename_in')
            if rc_download==True or is_file_fresh( filename_in, int(cfg.get('basic','срок годности'))):
                convert_excel2csv(cfg)



if __name__ == '__main__':
    myName = os.path.basename(os.path.splitext(sys.argv[0])[0])
    mydir    = os.path.dirname (sys.argv[0])
    print(mydir, myName)
    main( myName)
